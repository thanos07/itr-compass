"""Stateless tax-document parser for the ITR File web application.

The service performs structural text extraction only. It does not classify tax
items, calculate tax, or persist uploaded files. Ideas and portions are adapted
from the MIT-licensed prepare-india-tax-return generic parser.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import zipfile
from pathlib import PurePosixPath
from defusedxml import ElementTree as ET

from pypdf import PdfReader
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

MAX_TEXT = 1_000_000
MAX_ARCHIVE_MEMBERS = 100
MAX_EXPANDED_BYTES = int(os.getenv("MAX_EXPANDED_MB", "60")) * 1024 * 1024

app = FastAPI(title="ITR File parser", version="0.3.0")

origins = [origin.strip() for origin in os.getenv("ALLOWED_ORIGINS", "http://localhost:3000").split(",") if origin.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=False,
    allow_methods=["POST", "GET"],
    allow_headers=["*"],
)


def safe_name(name: str) -> str:
    return PurePosixPath(name.replace("\\", "/")).name or "upload"


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def parse_pdf(data: bytes, password: str | None) -> tuple[str, int, list[str]]:
    warnings: list[str] = []
    try:
        document = PdfReader(io.BytesIO(data), strict=False)
    except Exception as exc:
        raise HTTPException(400, f"Could not open PDF: {exc}") from exc
    if document.is_encrypted:
        try:
            unlocked = bool(password) and bool(document.decrypt(password or ""))
        except Exception as exc:
            raise HTTPException(422, "The PDF is encrypted and could not be decrypted with the supplied password.") from exc
        if not unlocked:
            raise HTTPException(422, "The PDF is password protected. Supply the correct document password.")
    chunks: list[str] = []
    total_length = 0
    has_embedded_text = False
    for index, page in enumerate(document.pages, start=1):
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""
            warnings.append(f"Page {index} could not be extracted and was left blank.")
        if page_text.strip():
            has_embedded_text = True
        chunk = f"[Page {index}]\n{page_text}"
        chunks.append(chunk)
        total_length += len(chunk)
        if total_length > MAX_TEXT:
            warnings.append("Text output was truncated at the service limit.")
            break
    if not has_embedded_text:
        warnings.append("No embedded text was found. This may be a scanned PDF and OCR is not included in the free parser.")
    return "\n\n".join(chunks)[:MAX_TEXT], len(document.pages), warnings


def parse_xlsx(data: bytes) -> tuple[str, int, list[str]]:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not open spreadsheet: {exc}") from exc
    chunks: list[str] = []
    row_count = 0
    warnings: list[str] = []
    for worksheet in workbook.worksheets:
        chunks.append(f"[Sheet: {worksheet.title}]")
        for row in worksheet.iter_rows(values_only=True):
            row_count += 1
            chunks.append(" | ".join("" if value is None else str(value) for value in row))
            if sum(map(len, chunks)) > MAX_TEXT:
                warnings.append("Spreadsheet output was truncated at the service limit.")
                return "\n".join(chunks)[:MAX_TEXT], row_count, warnings
    return "\n".join(chunks), row_count, warnings


def parse_docx(data: bytes) -> tuple[str, int, list[str]]:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            info = archive.getinfo("word/document.xml")
            if info.file_size > MAX_EXPANDED_BYTES:
                raise HTTPException(413, "DOCX expands beyond the service safety limit.")
            xml = archive.read("word/document.xml")
    except Exception as exc:
        raise HTTPException(400, f"Could not open DOCX: {exc}") from exc
    try:
        root = ET.fromstring(xml)
    except Exception as exc:
        raise HTTPException(400, "The DOCX XML is malformed or violates XML safety limits.") from exc
    texts = [node.text or "" for node in root.iter() if node.tag.endswith("}t")]
    return " ".join(texts)[:MAX_TEXT], 1, []


def parse_archive(data: bytes, password: str | None) -> tuple[str, int, list[str]]:
    warnings: list[str] = []
    chunks: list[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            infos = [info for info in archive.infolist() if not info.is_dir()]
            if sum(info.file_size for info in infos) > MAX_EXPANDED_BYTES:
                raise HTTPException(413, "ZIP expands beyond the service safety limit.")
            names = [info.filename for info in infos[:MAX_ARCHIVE_MEMBERS]]
            if len(infos) > MAX_ARCHIVE_MEMBERS:
                warnings.append("Only the first 100 archive members were read.")
            pwd = password.encode() if password else None
            for name in names:
                try:
                    member = archive.read(name, pwd=pwd)
                except RuntimeError:
                    raise HTTPException(422, "The ZIP is password protected. Supply the correct document password.")
                suffix = PurePosixPath(name).suffix.lower()
                if suffix in {".txt", ".csv", ".tsv", ".json", ".xml", ".html", ".md"}:
                    chunks.append(f"[Archive member: {name}]\n{decode_text(member)}")
                if sum(map(len, chunks)) > MAX_TEXT:
                    warnings.append("Archive output was truncated at the service limit.")
                    break
    except zipfile.BadZipFile as exc:
        raise HTTPException(400, "Invalid ZIP archive.") from exc
    return "\n\n".join(chunks)[:MAX_TEXT], len(names), warnings


def parse_data(data: bytes, name: str, password: str | None) -> tuple[str, int, list[str]]:
    lower = name.lower()
    if data.startswith(b"%PDF-") or lower.endswith(".pdf"):
        return parse_pdf(data, password)
    if lower.endswith((".xlsx", ".xlsm", ".xltx")):
        return parse_xlsx(data)
    if lower.endswith(".docx"):
        return parse_docx(data)
    if lower.endswith(".zip"):
        return parse_archive(data, password)
    if lower.endswith((".csv", ".tsv")):
        text = decode_text(data)
        delimiter = "\t" if lower.endswith(".tsv") else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        normalized = "\n".join(f"{index}: {' | '.join(row)}" for index, row in enumerate(rows[:50000], start=1))
        return normalized[:MAX_TEXT], len(rows), ["Delimited output was truncated."] if len(normalized) > MAX_TEXT else []
    if lower.endswith(".json"):
        try:
            payload = json.loads(decode_text(data))
        except json.JSONDecodeError as exc:
            raise HTTPException(400, f"Invalid JSON: {exc}") from exc
        text = json.dumps(payload, indent=2, ensure_ascii=False)
        return text[:MAX_TEXT], 1, ["JSON output was truncated."] if len(text) > MAX_TEXT else []
    if lower.endswith((".txt", ".md", ".html", ".xml", ".log")):
        text = decode_text(data)
        return text[:MAX_TEXT], text.count("\n") + 1, ["Text output was truncated."] if len(text) > MAX_TEXT else []
    raise HTTPException(415, "Unsupported file type. Use PDF, JSON, CSV/TSV, XLSX, DOCX, TXT/XML/HTML/MD, or ZIP.")


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/parse")
async def parse(
    file: UploadFile = File(...),
    password: str | None = Form(default=None),
    consent_acknowledged: bool = Form(default=False),
) -> dict[str, object]:
    if not consent_acknowledged:
        raise HTTPException(400, "Explicit processing consent acknowledgement is required for this parser request.")
    max_bytes = int(os.getenv("MAX_UPLOAD_MB", "20")) * 1024 * 1024
    data = await file.read(max_bytes + 1)
    if len(data) > max_bytes:
        raise HTTPException(413, f"File exceeds the {max_bytes // 1024 // 1024} MB limit.")
    name = safe_name(file.filename or "upload")
    text, units, warnings = parse_data(data, name, password)
    return {
        "name": name,
        "sha256": hashlib.sha256(data).hexdigest(),
        "text": text,
        "units": units,
        "warnings": warnings,
        "retained": False,
    }
